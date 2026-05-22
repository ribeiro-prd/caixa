from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

def write_analysis_excel(path: str | Path, sheets: dict[str, pd.DataFrame]) -> None:
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        for name, df in sheets.items():
            safe_name = name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
            ws = writer.book[safe_name]
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
            header_fill = PatternFill('solid', fgColor='1F4E78')
            header_font = Font(color='FFFFFF', bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            for col_idx, col in enumerate(ws.columns, start=1):
                max_len = max(len('' if cell.value is None else str(cell.value)) for cell in col)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
